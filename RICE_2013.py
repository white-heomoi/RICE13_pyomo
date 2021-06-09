import math
import pandas as pd
import pyomo.environ as pe
from Aux_func import model_res_to_dict, output_format, results_to_excel, coa_f, \
                     check_arg_T, check_arg_tstep, check_arg_tol, check_arg_max_iter, \
                     coa_to_analyse, check_bool_arg
from Coa_analysis import c_f_dif, int_st, dif_ext, ext_st, stab_c, coa_int, \
                         coa_ext
from openpyxl import load_workbook
import argparse

# Defining options to set number of time periods, precision of optimization solver,
# maximal number of iterations for the optimization algorithm
# and the cases to be analysed when calling the script from command line.
# If script is run from IDE, comment out tis section and uncomment the next one.

parser = argparse.ArgumentParser()
parser.add_argument('--T', default = 15, type = check_arg_T, help = 'Number of time periods to be considered (min = 2, max = 59)')

parser.add_argument('--tstep', default = 10, type = check_arg_tstep, help = 'Number of years between each time period (Accepted values: 1, 2, 5, 10, 20). Remember that tstep*T <= 590.')
                    
parser.add_argument('--tol', default = 7, type = check_arg_tol, help = 'Precision of the optimization algorithm expressed in number of decimal places (min = 7, max = 12)')

parser.add_argument('--max_it', default = 10000, type = check_arg_max_iter, help = 'Maximal number of iterations to be made by the optimization algorithm (min = 500, max = 25000)')
                    
parser.add_argument('--coop', default = "True", type = check_bool_arg, help = 'Establish if the full cooperative case is computed (True) or not (False): default is True')

parser.add_argument('--nc', default = "True", type = check_bool_arg, help = 'Establish if the non-cooperative case is computed (True) or not (False): default is True')
                    
parser.add_argument('--coalition', default = "none", type = str, help = 'Establish which intermediate coalition (1<|S|<N) should be computed. Available options are "none", default, "all" (takes long time) and a string with desired countries-regions: US, EU, JAP, RUS, EUR, CHI, IND, MEST, AFR, LAM, OHI, OTH')

args = parser.parse_args()

T = args.T
tstep = args.tstep
tol_a = args.tol
max_iter = args.max_it
if args.coop == 'True':
    coop_c = True
else:
    coop_c = False
if args.nc == 'True':
    nc_c = True
else:
    nc_c = False
coa_c = args.coalitions

if T*tstep > 590:
    from argparse import ArgumentTypeError
    raise ArgumentTypeError("The product of T and tstep must be lower than 590. When you choose a time step of 20 years, the maximal number of time periods (T) is 29.")

# =============================================================================
# If script is run from IDE, uncomment the following 5 lines and change, if you want, 
# their predefined values. Tips are provided above.
# 
# T = 15
# tstep = 10
# tol_a = 6
# max_iter = 10000
# coop_c = True
# nc_c = True
# coa_c = 'none' 
# 
# If run from command line, uncomment the following lines, whereas if the script is not
# run in whole but in "pieces" through some IDE, insert manually the path of the Data
# folder
# =============================================================================

tol = float('1e-'+str(tol_a))
 
from os import getcwd

data_path = getcwd()+'/Data/'

solver_path = getcwd()+'/Solver/ipopt.exe'

results_path = getcwd()+'/Results/'

# =============================================================================
# Uncomment and complete the following lines if the script is run from an IDE
# 
# data_path = INSERT DATA PATH HERE
# 
# solver_path = INSERT SOLVER PATH HERE 
# 
# results_path = INSERT PATH WHERE RESULTS SHOULD BE SAVED HERE
# 
# Import all parameters and time 0 values (plus TFP and Population growth that are pre-computed)
# =============================================================================

inv_p = pd.read_csv(data_path+"Time_countries_invariant_param.csv", sep=';', index_col=0)
c_var_p = pd.read_csv(data_path+"Countries_variant_param.csv", sep=';', index_col=0)
t0_v = pd.read_csv(data_path+"Time_0_values.csv", sep=';', index_col=0)
tfp_g = pd.read_csv(data_path+"TFP_growth.csv", sep=';', index_col=0)
pop_g = pd.read_csv(data_path+"Pop_growth.csv", sep=';', index_col=0)
s_r_t = pd.read_csv(data_path+"saving_rate_t.csv", sep=';', index_col=0)
slr_param = pd.read_csv(data_path+"SLR_time_variant_Param.csv", sep=';', index_col=0)

# Definition of Static Parameters (time and countries invariant)

b11     = float(inv_p.loc['b11'])          # atmosphere to atmosphere (b11)
b21     = float(inv_p.loc['b21'])          # biosphere/shallow oceans to atmosphere (b21)
b12     = float(inv_p.loc['b12'])          # atmosphere to biosphere/shallow oceans (b12)
b22     = float(inv_p.loc['b22'])          # biosphere/shallow oceans to biosphere/shallow oceans (b22)
b32     = float(inv_p.loc['b32'])          # deep oceans to biosphere/shallow oceans (b32)
b23     = float(inv_p.loc['b23'])          # biosphere/shallow oceans to deep oceans (b23)
b33     = float(inv_p.loc['b33'])          # deep oceans to deep oceans (b33)
c1      = float(inv_p.loc['c1'])           # Speed of adjustment parameter for atmospheric temperature
c2      = float(inv_p.loc['c2'])           # Transfer coefficient from previous time period (upper)
c3      = float(inv_p.loc['c3'])           # Coefficient of heat loss from atmosphere to oceans
c4      = float(inv_p.loc['c4'])           # Coefficient of heat gain by deep oceans
M_1900  = float(inv_p.loc['M_1900'])       # Reference concentration of CO2 (1900 level) 
eta     = float(inv_p.loc['eta'])          # Forcings of equilibrium CO2 doubling (Wm-2) 
t2xco2  = float(inv_p.loc['t2xco2'])       # Equilibrium temperature increase for CO2 doubling (t2xco2)
F_ex_1  = float(inv_p.loc['F_ex_1'])       # 2000 forcings other ghg
F_ex_2  = float(inv_p.loc['F_ex_2'])       # 2100 forcings other ghg
T_at_05 = float(inv_p.loc['T_at_05'])      # Initial atmospheric temperature, 2005 (deg. C above 1900)
T_at_15 = float(inv_p.loc['T_at_15'])      # Initial atmospheric temperature, 2015 (deg. C above 1900)
T_lo_0  = float(inv_p.loc['T_lo_0'])       # Initial temperature of deep oceans (deg. C above 1900)
M_at_0  = float(inv_p.loc['M_at_0'])       # Initial atmospheric concentration of CO2 (GTC, 2010)
M_up_0  = float(inv_p.loc['M_up_0'])       # Initial concentration of CO2 in biosphere/shallow oceans (GTC)
M_lo_0  = float(inv_p.loc['M_lo_0'])       # Initial concentration of CO2 in deep oceans (GTC)

 
# Definition of Static Parameters that are Country Dependent (time invariant)

gamma   = c_var_p.loc['gamma']            # Capital share
d_k     = c_var_p.loc['d_k']              # Rate of depreciation (percent per year)
rho     = c_var_p.loc['rho']              # Rate of social time preference
alpha   = c_var_p.loc['alpha']            # Marginal Utility of consumption
d_1    = c_var_p.loc['d_1']               # damage coefficient on temperature
d_2    = c_var_p.loc['d_2']               # damage coefficient on temperature squared
d_3    = c_var_p.loc['d_3']               # Exponent on damages
cat_t   = c_var_p.loc['cat_treash']       # Threshold for catastrophic damages
d_3_cat = c_var_p.loc['d_3_cat']          # Exponent for catastrophic damages
pback   = c_var_p.loc['pback']            # Price backstop technology (2005 US 000 $ per tC)
d_ab    = c_var_p.loc['d_ab']             # Decline of backstop price (per decade)
th_2    = c_var_p.loc['th_2']             # Coefficient on abatement level
E_0     = c_var_p.loc['E_0']              # CO2 emissions for 2005
Sig_0   = c_var_p.loc['Sig_0']            # Initial sigma (tC per $1000 GDP US $, 2005 prices) 2005
eland_0 = c_var_p.loc['Eland_0']          # Initial carbon emissions from land use change (GTC per year)
d_el    = c_var_p.loc['d_el']             # Decline rate of land emissions per decade
d_sig   = c_var_p.loc['d_sig']            # Decline of growth rate of sigma (sigma = emission intensity)
tr_sig  = c_var_p.loc['tr_sig']           # Trend of sigma (sigma = emission intensity)
tot_sig = c_var_p.loc['tot_sig']          # Total sigma growth (used to compute year 2015 of g_sigma)
sig_9506= c_var_p.loc['sig_9506']         # Average Sigma for period 95-06 
add_sig = c_var_p.loc['add_sig']          # Sigma increase for 2015 only
s_r     = c_var_p.loc['s_r']              # Saving rate (computed as average of all time periods in original RICE2013 excell file)
d1_slr  = c_var_p.loc['d1_slr']           # Damage coefficient for sea level rise (SLR) – linear
d2_slr  = c_var_p.loc['d2_slr']           # Damage coefficient for sea level rise (SLR) – quadratic
sig_15_add  = c_var_p.loc['sig_15_add']   # Multiplicative factor to add to sigma only for year 2015

# Values of variables at time zero 

Y_0     = t0_v.loc['Y_0']           # Initial output (2005 US International $, trillions)
rho_0   = t0_v.loc['rho_0']         # Initial real interest rate (percent per decade annualized)
K_0     = t0_v.loc['K_0']           # Initial K
A_0     = t0_v.loc['A_0']           # Calculated A (Initial)
s_0     = t0_v.loc['s_0']           # Initial savings rate
L_0     = t0_v.loc['L_0']           # Initial population
E_0     = t0_v.loc['E_0']           # Initial emissions

# TFP and population growth for all years (g_A(t) and g_L(t) in the equations of the paper)

g_A = tfp_g             # Total Factor Productivity (TFP = A) growth rate 
g_L = pop_g             # Population\labour (L) growth rate

# SLR damage parameter

slr_p = pd.Series(slr_param.loc['SLR_p'])

# Saving rate (Sig_I_t)) for all years. The yearly saving rate is taken from the 
# Excell version of RICE2013 solved for the base scenario. Sig_I_t varies yearly, 
# whereas Sig_I (s_r in the data file) is constant and is computed as the average of Sig_I_t.

Sig_I_t = s_r_t
for i in range(len(Sig_I_t.columns)):
    Sig_I_t.rename(columns={Sig_I_t.columns[i]:i}, inplace=True)

# Definition of Parameter values for all time periods

years = [2015+tstep*i for i in range(T)]
y_as_int = [i for i in range(T)]
countries = g_L.index
N = len(countries)

# Add missing years to g_L(t) (growth equal to 0 for all countries in long period)
mis_ind = [i for i in range(len(g_L.loc['US']),60)]
mis_dat = [0 for i in range(len(g_L.loc['US']),60)]
mis_g_L_dat = {i:mis_dat for i in g_L.index}
mis_g_L = pd.DataFrame(data = mis_g_L_dat, index=mis_ind)

g_L = g_L.T.append(mis_g_L, ignore_index=True).T
for i in range(len(g_A.columns)):
    g_A.rename(columns={g_A.columns[i]:i}, inplace=True)

for i in range(len(slr_p.index)):
    slr_p.rename(index={slr_p.index[i]:i}, inplace=True)

# Population (L(i,t))
L = pd.DataFrame(0, index=countries,columns=y_as_int)
for i in countries:
    for j in range(T):
        if j==0:
            L.loc[i,j] = L_0.loc[i]
        elif j==1:
            L.loc[i,j] = L.loc[i,j-1]*math.exp(g_L.loc[i,j]*10)
        else:
            if tstep==10:
                L.loc[i,j] = L.loc[i,j-1]*math.exp(g_L.loc[i,j]*tstep)
            elif tstep==20:
                L.loc[i,j] = L.loc[i,j-1]*math.exp(g_L.loc[i,j*2]*tstep)
            else:
                L.loc[i,j] = L.loc[i,j-1]*math.exp(g_L.loc[i,j//int(10/tstep)+1]*tstep)
        
# TFP (A(i,t))    
A = pd.DataFrame(0, index=countries,columns=y_as_int)
for i in countries:
    for j in range(T):
        if j==0:
            A.loc[i,j] = A_0.loc[i]
        elif j==1:
            A.loc[i,j] = A_0.loc[i]*math.exp(g_A.loc[i,j]*10)
        else:
            if tstep == 10:
                A.loc[i,j] = A.loc[i,j-1]*math.exp(g_A.loc[i,j]*tstep)
            elif tstep == 20:
                A.loc[i,j] = A.loc[i,j-1]*math.exp(g_A.loc[i,j*2]*tstep)
            else:
                A.loc[i,j] = A.loc[i,j-1]*math.exp(g_A.loc[i,j//int(10/tstep)+1]*tstep)

# Sigma growth (GSIG(i,t))
if tstep !=20:
    g_sig = pd.DataFrame(0, index=countries,columns=y_as_int)
else:
    g_sig = pd.DataFrame(0, index=countries,columns=[i for i in range(T*2)])
for i in countries:
    for j in range(T):
        if j==0:
            g_sig.loc[i,j] = 0 #tot_sig.loc[i]
        elif j==1:
            g_sig.loc[i,j] = tot_sig.loc[i]
        else:
            g_sig.loc[i,j] = tr_sig.loc[i]+(g_sig.loc[i,j-1]-tr_sig.loc[i])*(1-d_sig.loc[i])

# Sigma (SIGMA(i,t))
sig = pd.DataFrame(0, index=countries,columns=y_as_int)
for i in countries:
    for j in range(T):
        if j==0:
            sig.loc[i,j] = Sig_0.loc[i]
        elif j ==1:
            sig.loc[i,j] = sig.loc[i,j-1]*math.exp(g_sig.loc[i,j]*10)*sig_15_add[i]
        else:
            if tstep==10:
                sig.loc[i,j] = sig.loc[i,j-1]*math.exp(g_sig.loc[i,j]*tstep)
            elif tstep==20:
                sig.loc[i,j] = sig.loc[i,j-1]*math.exp(g_sig.loc[i,j*2]*tstep)
            else:
                sig.loc[i,j] = sig.loc[i,j-1]*math.exp(g_sig.loc[i,j//int(10/tstep)+1]*tstep)

# Emissions from land (ETREE(i,t)) - exogenous
eland = pd.DataFrame(0, index=countries, columns=y_as_int)
for i in countries:
    for j in range(T):
        if j==0:
            eland.loc[i,j] = eland_0.loc[i]
        elif j ==1:
            eland.loc[i,j] = eland.loc[i,j-1]*(1-d_el.loc[i])
        else:
            eland.loc[i,j] = eland.loc[i,j-1]*(1-d_el.loc[i]*(tstep/10))

# Exogenous radiative forcing (FORCOTH(t)) 
f_ex = pd.Series([F_ex_1+0.1*(F_ex_2-F_ex_1)*j if j<11 else F_ex_1+0.36 for j in range(T)], index=y_as_int)

if tstep != 10:   
    from sklearn.linear_model import LinearRegression
    from numpy import array as npa
    y_var = npa(f_ex[:min(11,T)])
    X_var = npa([i for i in range(min(11,T))])
    X_var = X_var.reshape(-1,1)
    mod = LinearRegression()
    mod.fit(X_var, y_var)
    switch_p = 10*10/tstep
    f_ex = [mod.intercept_ + mod.coef_[0]*i*(tstep/10) if i <= switch_p else 0.3 for i in range(1,T+1)]
    f_ex.insert(0, -0.06)

# Backstop price (used to compute COST1(i,t))
backstpr1 = pd.DataFrame(0, index=countries, columns=[i for i in range(60)])
for i in countries:
    for j in range(60):
        if j==0:
            backstpr1.loc[i,j] = pback.loc[i]
        elif j==1:
            backstpr1.loc[i,j] = pback.loc[i]*0.1+0.9*pback.loc[i]*(1-d_ab.loc[i])
        elif 2015 + j*tstep >= 2250:    # Assumption of a completely green technology cheaper than fossil fuels 
            backstpr1.loc[i,j] = 0
        else:
            backstpr1.loc[i,j] = pback.loc[i]*0.1+(backstpr1.loc[i,j-1]-0.1*pback.loc[i])*(1-d_ab.loc[i])

if tstep == 10:
    backstpr = backstpr1.loc[:,:T+1]    
elif tstep == 20:
    backstpr = pd.DataFrame(0, index=countries, columns=y_as_int)
    backstpr.loc[:, 0] = backstpr1.loc[:,0]
    for i in range(T):
        backstpr.loc[:, i+1]  = backstpr1.loc[:,i*2+1]   
else:
    backstpr = pd.DataFrame(0, index=countries, columns=y_as_int)
    backstpr.loc[:, 0] = backstpr1.loc[:,0]
    n_per = int(10/tstep)
    n_T = int(T/n_per)
    for i in range(1, n_T+1):
        for j in range(n_per):
            backstpr.loc[:, (i-1)*n_per + j +1] =  backstpr1.loc[:,i] + \
                        j*(backstpr1.loc[:,i+1] - backstpr1.loc[:,i])/n_per

# Cost coefficient for abatement (COST1(i,t)) that varies during time
theta1 = pd.DataFrame(0, index=countries, columns=y_as_int)
for i in countries:
    for j in range(T):
        theta1.loc[i,j] = backstpr.loc[i,j]*sig.loc[i,j]/th_2.loc[i]
        
# Investment at time 0 (I_0[i])
I_0 = Y_0*Sig_I_t.loc[:,0]

# Pyomo model

m = pe.ConcreteModel()

# Sets declaration

m.t = pe.Set(initialize=[i for i in range(1,T)])      # Time periods
m.mC = pe.Set(initialize=countries)                   # Countries

t = [i for i in range(1,T)]
mC = countries

# Loading variables starting values (Variables are defined and explained later)

U_init = pd.read_csv(data_path+'U_init.csv', sep=';', index_col=0)
K_init = pd.read_csv(data_path+'K_init.csv', sep=';', index_col=0)
I_init = pd.read_csv(data_path+'I_init.csv', sep=';', index_col=0)
Q_init = pd.read_csv(data_path+'Q_init.csv', sep=';', index_col=0)
Y_init = pd.read_csv(data_path+'Y_init.csv', sep=';', index_col=0)
mu_init = pd.read_csv(data_path+'mu_init.csv', sep=';', index_col=0)
AB_init = pd.read_csv(data_path+'AB_init.csv', sep=';', index_col=0)
D_init = pd.read_csv(data_path+'D_init.csv', sep=';', index_col=0)
C_init = pd.read_csv(data_path+'C_init.csv', sep=';', index_col=0)
E_ind_init = pd.read_csv(data_path+'E_ind_init.csv', sep=';', index_col=0)
oth_var = pd.read_csv(data_path+'Var_country_independent_init.csv', sep=';', index_col=0)

# Keep only relevant number of time periods (max T = 59)
# Do not know why, but it imports the columns indexes as str instead of int (need to convert them)

U_init = pd.read_csv(data_path+'U_init.csv', sep=';', index_col=0)
U_init.columns = U_init.columns.astype(int)
K_init = pd.read_csv(data_path+'K_init.csv', sep=';', index_col=0)
K_init.columns = K_init.columns.astype(int)
I_init = pd.read_csv(data_path+'I_init.csv', sep=';', index_col=0)
I_init.columns = I_init.columns.astype(int)
Q_init = pd.read_csv(data_path+'Q_init.csv', sep=';', index_col=0)
Q_init.columns = Q_init.columns.astype(int)
Y_init = pd.read_csv(data_path+'Y_init.csv', sep=';', index_col=0)
Y_init.columns = Y_init.columns.astype(int)
mu_init = pd.read_csv(data_path+'mu_init.csv', sep=';', index_col=0)
mu_init.columns = mu_init.columns.astype(int)
AB_init = pd.read_csv(data_path+'AB_init.csv', sep=';', index_col=0)
AB_init.columns = AB_init.columns.astype(int)
D_init = pd.read_csv(data_path+'D_init.csv', sep=';', index_col=0)
D_init.columns = D_init.columns.astype(int)
C_init = pd.read_csv(data_path+'C_init.csv', sep=';', index_col=0)
C_init.columns = C_init.columns.astype(int)
E_ind_init = pd.read_csv(data_path+'E_ind_init.csv', sep=';', index_col=0)
E_ind_init.columns = E_ind_init.columns.astype(int)
oth_var = pd.read_csv(data_path+'Var_country_independent_init.csv', sep=';', index_col=0)
oth_var.columns = oth_var.columns.astype(int)

# Keep only relevant number of time periods (max T = 59)
# Do not know why, but it imports the columns indexes as str instead of int (need to convert them)

if tstep == 10:
    U_init = U_init[U_init.columns[1:T+1]]
    K_init = K_init[K_init.columns[1:T+1]]
    I_init = I_init[I_init.columns[1:T+1]]
    Q_init = Q_init[Q_init.columns[1:T+1]]
    Y_init = Y_init[Y_init.columns[1:T+1]]
    mu_init = mu_init[mu_init.columns[1:T+1]]
    AB_init = AB_init[AB_init.columns[1:T+1]]
    D_init = D_init[D_init.columns[1:T+1]]
    C_init = C_init[C_init.columns[1:T+1]]
    E_ind_init = E_ind_init[E_ind_init.columns[1:T+1]]
    E_tot_init = pd.Series(oth_var.loc['E_tot'][oth_var.loc['E_tot'].index[1:T+1]])
    M_at_init = pd.Series(oth_var.loc['M_at'][oth_var.loc['M_at'].index[1:T+1]])
    M_up_init = pd.Series(oth_var.loc['M_up'][oth_var.loc['M_up'].index[1:T+1]])
    M_lo_init = pd.Series(oth_var.loc['M_lo'][oth_var.loc['M_lo'].index[1:T+1]])
    T_at_init = pd.Series(oth_var.loc['T_at'][oth_var.loc['T_at'].index[1:T+1]])
    T_lo_init = pd.Series(oth_var.loc['T_lo'][oth_var.loc['T_lo'].index[1:T+1]])
    F_init = pd.Series(oth_var.loc['F'][oth_var.loc['F'].index[1:T+1]])    
elif tstep == 20:
    col_ind = [i for i in range(1,T*2+1,2)]
    col_new_ind = [i for i in range(1,T+1)]
    U_init = U_init[U_init.columns[col_ind]]
    U_init.columns = col_new_ind
    K_init = K_init[K_init.columns[col_ind]]
    K_init.columns = col_new_ind
    I_init = I_init[I_init.columns[col_ind]]
    I_init.columns = col_new_ind
    Q_init = Q_init[Q_init.columns[col_ind]]
    Q_init.columns = col_new_ind
    Y_init = Y_init[Y_init.columns[col_ind]]
    Y_init.columns = col_new_ind
    mu_init = mu_init[mu_init.columns[col_ind]]
    mu_init.columns = col_new_ind
    AB_init = AB_init[AB_init.columns[col_ind]]
    AB_init.columns = col_new_ind
    D_init = D_init[D_init.columns[col_ind]]
    D_init.columns = col_new_ind
    C_init = C_init[C_init.columns[col_ind]]
    C_init.columns = col_new_ind
    E_ind_init = E_ind_init[E_ind_init.columns[col_ind]]
    E_ind_init.columns = col_new_ind
    E_tot_init = pd.Series(oth_var.loc['E_tot'][oth_var.loc['E_tot'].index[col_ind]])
    E_tot_init.index = col_new_ind
    M_at_init = pd.Series(oth_var.loc['M_at'][oth_var.loc['M_at'].index[col_ind]])
    M_at_init.index = col_new_ind
    M_up_init = pd.Series(oth_var.loc['M_up'][oth_var.loc['M_up'].index[col_ind]])
    M_up_init.index = col_new_ind
    M_lo_init = pd.Series(oth_var.loc['M_lo'][oth_var.loc['M_lo'].index[col_ind]])
    M_lo_init.index = col_new_ind
    T_at_init = pd.Series(oth_var.loc['T_at'][oth_var.loc['T_at'].index[col_ind]])
    T_at_init.index = col_new_ind
    T_lo_init = pd.Series(oth_var.loc['T_lo'][oth_var.loc['T_lo'].index[col_ind]])
    T_lo_init.index = col_new_ind
    F_init = pd.Series(oth_var.loc['F'][oth_var.loc['F'].index[col_ind]])
    F_init.index = col_new_ind
else:
    U_init0 = U_init[U_init.columns[1:T+1]]
    U_init = U_init[U_init.columns[1:T+1]]
    K_init0 = K_init[K_init.columns[1:T+1]]
    K_init = K_init[K_init.columns[1:T+1]]
    I_init0 = I_init[I_init.columns[1:T+1]]
    I_init = I_init[I_init.columns[1:T+1]]
    Q_init0 = Q_init[Q_init.columns[1:T+1]]
    Q_init = Q_init[Q_init.columns[1:T+1]]
    Y_init0 = Y_init[Y_init.columns[1:T+1]]
    Y_init = Y_init[Y_init.columns[1:T+1]]
    mu_init0 = mu_init[mu_init.columns[1:T+1]]
    mu_init = mu_init[mu_init.columns[1:T+1]]
    AB_init0 = AB_init[AB_init.columns[1:T+1]]
    AB_init = AB_init[AB_init.columns[1:T+1]]
    D_init0 = D_init[D_init.columns[1:T+1]]
    D_init = D_init[D_init.columns[1:T+1]]
    C_init0 = C_init[C_init.columns[1:T+1]]
    C_init = C_init[C_init.columns[1:T+1]]
    E_ind_init0 = E_ind_init[E_ind_init.columns[1:T+1]]
    E_ind_init = E_ind_init[E_ind_init.columns[1:T+1]]
    E_tot_init0 = pd.Series(oth_var.loc['E_tot'][oth_var.loc['E_tot'].index[1:T+1]])
    M_at_init0 = pd.Series(oth_var.loc['M_at'][oth_var.loc['M_at'].index[1:T+1]])
    M_up_init0 = pd.Series(oth_var.loc['M_up'][oth_var.loc['M_up'].index[1:T+1]])
    M_lo_init0 = pd.Series(oth_var.loc['M_lo'][oth_var.loc['M_lo'].index[1:T+1]])
    T_at_init0 = pd.Series(oth_var.loc['T_at'][oth_var.loc['T_at'].index[1:T+1]])
    T_lo_init0 = pd.Series(oth_var.loc['T_lo'][oth_var.loc['T_lo'].index[1:T+1]])
    F_init0 = pd.Series(oth_var.loc['F'][oth_var.loc['F'].index[1:T+1]]) 
    E_tot_init = pd.Series(oth_var.loc['E_tot'][oth_var.loc['E_tot'].index[1:T+1]])
    M_at_init = pd.Series(oth_var.loc['M_at'][oth_var.loc['M_at'].index[1:T+1]])
    M_up_init = pd.Series(oth_var.loc['M_up'][oth_var.loc['M_up'].index[1:T+1]])
    M_lo_init = pd.Series(oth_var.loc['M_lo'][oth_var.loc['M_lo'].index[1:T+1]])
    T_at_init = pd.Series(oth_var.loc['T_at'][oth_var.loc['T_at'].index[1:T+1]])
    T_lo_init = pd.Series(oth_var.loc['T_lo'][oth_var.loc['T_lo'].index[1:T+1]])
    F_init = pd.Series(oth_var.loc['F'][oth_var.loc['F'].index[1:T+1]])

    n_per = int(10/tstep)
    n_T = int(T/n_per)
    for i in range(n_T):
        for j in range(n_per):
            U_init.loc[:, (i)*n_per + j +1] = U_init0.loc[:,i+1] + j*(U_init0.loc[:,i+2] - U_init0.loc[:,i+1])/n_per
            K_init.loc[:, (i)*n_per + j +1] = K_init0.loc[:,i+1] + j*(K_init0.loc[:,i+2] - K_init0.loc[:,i+1])/n_per
            I_init.loc[:, (i)*n_per + j +1] = I_init0.loc[:,i+1] + j*(I_init0.loc[:,i+2] - I_init0.loc[:,i+1])/n_per
            Q_init.loc[:, (i)*n_per + j +1] = Q_init0.loc[:,i+1] + j*(Q_init0.loc[:,i+2] - Q_init0.loc[:,i+1])/n_per
            Y_init.loc[:, (i)*n_per + j +1] = Y_init0.loc[:,i+1] + j*(Y_init0.loc[:,i+2] - Y_init0.loc[:,i+1])/n_per
            mu_init.loc[:, (i)*n_per + j +1] = mu_init0.loc[:,i+1] + j*(mu_init0.loc[:,i+2] - mu_init0.loc[:,i+1])/n_per
            AB_init.loc[:, (i)*n_per + j +1] = AB_init0.loc[:,i+1] + j*(AB_init0.loc[:,i+2] - AB_init0.loc[:,i+1])/n_per
            D_init.loc[:, (i)*n_per + j +1] = D_init0.loc[:,i+1] + j*(D_init0.loc[:,i+2] - D_init0.loc[:,i+1])/n_per
            C_init.loc[:, (i)*n_per + j +1] = C_init0.loc[:,i+1] + j*(C_init0.loc[:,i+2] - C_init0.loc[:,i+1])/n_per
            E_ind_init.loc[:, (i)*n_per + j +1] = E_ind_init0.loc[:,i+1] + j*(E_ind_init0.loc[:,i+2] - E_ind_init0.loc[:,i+1])/n_per
            E_tot_init.loc[(i)*n_per + j +1] = E_tot_init0.loc[i+1] + j*(E_tot_init0.loc[i+2] - E_tot_init0.loc[i+1])/n_per
            M_at_init.loc[(i)*n_per + j +1] = M_at_init0.loc[i+1] + j*(M_at_init0.loc[i+2] - M_at_init0.loc[i+1])/n_per
            M_up_init.loc[(i)*n_per + j +1] = M_up_init0.loc[i+1] + j*(M_up_init0.loc[i+2] - M_up_init0.loc[i+1])/n_per
            M_lo_init.loc[(i)*n_per + j +1] = M_lo_init0.loc[i+1] + j*(M_lo_init0.loc[i+2] - M_lo_init0.loc[i+1])/n_per
            T_at_init.loc[(i)*n_per + j +1] = T_at_init0.loc[i+1] + j*(T_at_init0.loc[i+2] - T_at_init0.loc[i+1])/n_per
            T_lo_init.loc[(i)*n_per + j +1] = T_lo_init0.loc[i+1] + j*(T_lo_init0.loc[i+2] - T_lo_init0.loc[i+1])/n_per
            F_init.loc[(i)*n_per + j +1] = F_init0.loc[i+1] + j*(F_init0.loc[i+2] - F_init0.loc[i+1])/n_per

S_init = I_init/Y_init

# Functions to initialize variables (give them starting values)

def U_init_f(m, mC, t):
    return U_init.loc[mC, t]

def K_init_f(m, mC, t):
    return K_init.loc[mC, t]

def S_init_f(m, mC, t):
    return S_init.loc[mC, t]

def I_init_f(m, mC, t):
    return I_init.loc[mC, t]

def Q_init_f(m, mC, t):
    return Q_init.loc[mC, t]

def Y_init_f(m, mC, t):
    return Y_init.loc[mC, t]

def mu_init_f(m, mC, t):
    return mu_init.loc[mC, t]

def AB_init_f(m, mC, t):
    return AB_init.loc[mC, t]

def D_init_f(m, mC, t):
    return D_init.loc[mC, t]

def C_init_f(m, mC, t):
    return C_init.loc[mC, t]

def E_ind_init_f(m, mC, t):
    return E_ind_init.loc[mC, t]

def E_tot_init_f(m, t):
    return E_tot_init[t]

def M_at_init_f(m, t):
    return M_at_init[t]

def M_up_init_f(m, t):
    return M_up_init[t]

def M_lo_init_f(m, t):
    return M_lo_init[t]

def T_at_init_f(m, t):
    return T_at_init[t]

def T_lo_init_f(m, t):
    return T_lo_init[t]

def F_init_f(m, t):
    return F_init[t]

# Variables definition

m.U = pe.Var(m.mC, m.t, domain=pe.Reals, initialize = U_init_f)                                       # Utility (at single point in time)
m.K = pe.Var(m.mC, m.t, domain=pe.NonNegativeReals, bounds=(1,float('inf')), initialize = K_init_f)   # Capital
m.S = pe.Var(m.mC, m.t, domain=pe.NonNegativeReals, bounds=(0,1), initialize = S_init_f)              # Saving rate (as control variable)
m.I = pe.Var(m.mC, m.t, domain=pe.NonNegativeReals, initialize = I_init_f)                            # Investments (= savings)
m.Q = pe.Var(m.mC, m.t, domain=pe.NonNegativeReals, initialize = Q_init_f)                            # Gross output
m.Y = pe.Var(m.mC, m.t, domain=pe.NonNegativeReals, initialize = Y_init_f)                            # Net output (after damages and expenditures for abatement)
m.mu = pe.Var(m.mC, m.t, domain=pe.NonNegativeReals, bounds=(0,1), initialize = mu_init_f)            # Abatement rate (control variable)
m.AB = pe.Var(m.mC, m.t, domain=pe.NonNegativeReals, bounds=(0,1), initialize = AB_init_f)            # Abatement costs (proportional)
m.D = pe.Var(m.mC, m.t, domain=pe.NonNegativeReals, initialize = D_init_f)                            # Environmental damages
m.C = pe.Var(m.mC, m.t, domain=pe.NonNegativeReals, initialize = C_init_f)                            # Consumption
m.E_ind = pe.Var(m.mC, m.t, domain=pe.NonNegativeReals, initialize = E_ind_init_f)                    # Industrial emissions
m.E_tot = pe.Var(m.t, domain=pe.NonNegativeReals, initialize = E_tot_init_f)                          # Total global emissions
m.M_at = pe.Var(m.t, domain=pe.Reals , bounds=(0.001, float('inf')), initialize = M_at_init_f)        # Atmospheric GHG concentration
m.M_up = pe.Var(m.t, domain=pe.Reals, initialize = M_up_init_f)                                       # Ocean (upper stratum) GHG concentration
m.M_lo = pe.Var(m.t, domain=pe.Reals, initialize = M_lo_init_f)                                       # Ocean (lower stratum) GHG concentration
m.T_at = pe.Var(m.t, domain=pe.Reals, initialize = T_at_init_f)                                       # Atmospheric temperature (increase over normal)
m.T_lo = pe.Var(m.t, domain=pe.Reals, initialize = T_lo_init_f)                                       # Temperature (lower ocean)
m.F = pe.Var(m.t, domain=pe.Reals, initialize = F_init_f)                                             # Radiative force

# Equations definition

# Utility
def U_eq(m, mC, t):
    return m.U[mC, t] == L.loc[mC, t]*((1/(1-alpha[mC]))*((m.C[mC, t]/L.loc[mC, t])*1000)**(1-alpha[mC])+1)/(1+rho[mC])**(tstep*t)
m.U_eq = pe.Constraint(m.mC, m.t, rule=U_eq)

#Capital
def K_eq(m, mC, t):
    if t == 1:
        return m.K[mC, t] == tstep*I_0[mC] + K_0[mC]*(1-d_k[mC])**tstep
    else:
        return m.K[mC, t] == tstep*m.I[mC, t-1] + m.K[mC, t-1]*(1-d_k[mC])**tstep
m.K_eq = pe.Constraint(m.mC, m.t, rule=K_eq)

# =============================================================================
# Investment (as a fixed proportion of net output)
# def I_eq(m, mC, t):
#     return m.I[mC, t] == Sig_I_t.loc[mC, t]*m.Y[mC, t]
# m.I_eq = pe.Constraint(m.mC, m.t, rule=I_eq)
# =============================================================================

# Investments as a variable (driven by saving rate that is a control variable)
def I_eq(m, mC, t):
    return m.I[mC, t] == m.S[mC, t]*m.Y[mC, t]
m.I_eq = pe.Constraint(m.mC, m.t, rule=I_eq)

# Gross output
def Q_eq(m, mC, t):
    return m.Q[mC, t] == A.loc[mC, t]*m.K[mC, t]**gamma[mC]*(L.loc[mC, t]/1000)**(1-gamma[mC])
m.Q_eq = pe.Constraint(mC, t, rule=Q_eq)

# Net output (after abatement costs and environmental damages)
def Y_eq(m, mC, t):
    return m.Y[mC, t] == m.Q[mC, t] - m.D[mC, t]*m.Q[mC, t]/(1+m.D[mC, t]**tstep) - m.AB[mC, t]
m.Y_eq = pe.Constraint(m.mC, m.t, rule=Y_eq)

# Abatement costs as proportion of output
def AB_eq(m, mC, t):
    return m.AB[mC, t] == (theta1.loc[mC, t]*m.mu[mC, t]**th_2[mC])*m.Q[mC, t]
m.AB_eq = pe.Constraint(m.mC, m.t, rule=AB_eq)

# Environmental damages (including the ones caused by sea level rise)
def D_eq(m, mC, t):
    return m.D[mC, t] == (d_1[mC]*m.T_at[t] + d_2[mC]*m.T_at[t]**d_3[mC])*0.01 \
                         + 2*(d1_slr[mC]*slr_p[t-1] + d2_slr[mC]*slr_p[t-1]**2)*(m.Q[mC, t]/Y_0[mC])**0.25
m.D_eq = pe.Constraint(m.mC, m.t, rule=D_eq)

# Consumption
def C_eq(m, mC, t):
    return m.C[mC, t] == m.Y[mC, t] - m.I[mC, t]
m.C_eq = pe.Constraint(m.mC, m.t, rule=C_eq)

# Industrial emissions
def E_ind_eq(m, mC, t):
    return m.E_ind[mC, t] == sig.loc[mC, t]*((1-m.mu[mC, t])*m.Q[mC, t])
m.E_ind_eq = pe.Constraint(m.mC, m.t, rule=E_ind_eq)

# Total emissions including emissions from LUC (summed over all countries)
def E_tot_eq(m, t):
    return m.E_tot[t] == sum(m.E_ind[i, t] + eland.loc[i, t] for i in mC)
m.E_tot_eq = pe.Constraint(m.t, rule=E_tot_eq)

# Athmospheric GHG concentration
def M_at_eq(m, t):
    if t == 1:
        return m.M_at[t] == (sum(E_0[i] + eland_0[i] for i in m.mC))*tstep + b11*M_at_0 + b21*M_up_0
    else:
        return m.M_at[t] == m.E_tot[t-1]*tstep + b11*m.M_at[t-1] + b21*m.M_up[t-1]
m.M_at_eq = pe.Constraint(m.t, rule=M_at_eq)

# Biosphere and upper ocean GHG concentration
def M_up_eq(m, t):
    if t == 1:
        return m.M_up[t] ==  b12*M_at_0 + b22*M_up_0 + b32*M_lo_0
    else:
        return m.M_up[t] == b12*m.M_at[t-1] + b22*m.M_up[t-1] + b32*m.M_lo[t-1]
m.M_up_eq = pe.Constraint(m.t, rule=M_up_eq)

# Lower ocean GHG concentration
def M_lo_eq(m, t):
    if t == 1:
        return m.M_lo[t] ==  b23*M_up_0 + b33*M_lo_0
    else:
        return m.M_lo[t] == b23*m.M_up[t-1] + b33*m.M_lo[t-1]
m.M_lo_eq = pe.Constraint(m.t, rule=M_lo_eq)

# Athmospheric temperature increase
def T_at_eq(m, t):
    if t == 1:
        return m.T_at[t] == T_at_15
    else:
        return m.T_at[t] == m.T_at[t-1] + c1*(m.F[t] - c2*m.T_at[t-1] - c3*(m.T_at[t-1] - m.T_lo[t-1]))
m.T_at_eq = pe.Constraint(m.t, rule=T_at_eq)

# Lower strata temperature increase
def T_lo_eq(m, t):
    if t == 1:
        return m.T_lo[t] ==  T_lo_0 + c4*(T_at_05 - T_lo_0)
    else:
        return m.T_lo[t] == m.T_lo[t-1] + c4*(m.T_at[t-1] - m.T_lo[t-1]) 
m.T_lo_eq = pe.Constraint(m.t, rule=T_lo_eq)

# Radiative force
def F_eq(m, t):
    return m.F[t] == eta*pe.log(m.M_at[t]/M_1900)/pe.log(2) + f_ex[t]
m.F_eq = pe.Constraint(m.t, rule=F_eq)

# Transform dataset for variable initialization to dictionary that is required 
# to use the var.set_values() method to reassign values to pyomo variables.
# Using the values of the last computed solution rather than the values of the
# initialization seems to cause the algorithm not to converge. Maybe with a 
# different solver from IPOPT this problem would not rise. 

U_init_dic = {(i, j): U_init.loc[i][j] for i in U_init.index for j in U_init.columns[:-1]}
K_init_dic = {(i, j): K_init.loc[i][j] for i in K_init.index for j in K_init.columns[:-1]}
S_init_dic = {(i, j): S_init.loc[i][j] for i in S_init.index for j in S_init.columns[:-1]}
I_init_dic = {(i, j): I_init.loc[i][j] for i in I_init.index for j in I_init.columns[:-1]}
Y_init_dic = {(i, j): Y_init.loc[i][j] for i in Y_init.index for j in Y_init.columns[:-1]}
Q_init_dic = {(i, j): Q_init.loc[i][j] for i in Q_init.index for j in Q_init.columns[:-1]}
AB_init_dic = {(i, j): AB_init.loc[i][j] for i in AB_init.index for j in AB_init.columns[:-1]}
I_init_dic = {(i, j): I_init.loc[i][j] for i in I_init.index for j in I_init.columns[:-1]}
D_init_dic = {(i, j): D_init.loc[i][j] for i in D_init.index for j in D_init.columns[:-1]}
C_init_dic = {(i, j): C_init.loc[i][j] for i in C_init.index for j in C_init.columns[:-1]}
E_ind_init_dic = {(i, j): E_ind_init.loc[i][j] for i in E_ind_init.index for j in E_ind_init.columns[:-1]}
E_tot_init_dic = {int(j): E_tot_init.loc[j] for j in E_tot_init.index[:-1]}
M_at_init_dic = {int(j): M_at_init.loc[j] for j in M_at_init.index[:-1]}
M_lo_init_dic = {int(j): M_lo_init.loc[j] for j in M_lo_init.index[:-1]}
M_up_init_dic = {int(j): M_up_init.loc[j] for j in M_up_init.index[:-1]}
T_at_init_dic = {int(j): T_at_init.loc[j] for j in T_at_init.index[:-1]}
T_lo_init_dic = {int(j): T_lo_init.loc[j] for j in T_lo_init.index[:-1]}
F_init_dic = {int(j): F_init.loc[j] for j in F_init.index[:-1]}

# Setting optimization algorithm specifications
opt = pe.SolverFactory('ipopt', executable=solver_path)
opt.options['max_iter']= max_iter
opt.options['tol'] = tol

if coop_c or coa_c == 'all':
    # Objective function (for the cooperative case)
    def obj_eq(m):
        return sum(m.U[i, j] for i in m.mC for j in m.t)
    m.obj = pe.Objective(rule=obj_eq, sense=pe.maximize) 
    
    # Solving the optimization problem
    opt.solve(m)
    
    # Delete the objective function for other eventual optimizations
    del m.obj

    ### Cooperative results ###
    
    # Data Frame of results for full cooperative case
    
    coop_dict = model_res_to_dict(m)
    
    res_coop = output_format(countries, coop_dict, t, T)

    coop_TU = [sum(res_coop[i].loc['U',:]) for i in countries]
    
    #Save output on Excell file
    
    results_to_excel(res_coop, countries, results_path, 'coop.xlsx')
    
    # Stability Analysis
    # (For full cooperation, only internal stability has to be computed)
if coop_c and coa_c != 'all':
    gr_coa = [1 for i in range(N)]
    l_int_c = coa_int(gr_coa)
    coa_res = []
    for i in l_int_c:
        m.U.set_values(U_init_dic)                                       
        m.I.set_values(I_init_dic)
        m.S.set_values(S_init_dic)                            
        m.Q.set_values(Q_init_dic)                           
        m.Y.set_values(Y_init_dic)                                       
        m.AB.set_values(AB_init_dic)            
        m.D.set_values(D_init_dic)                            
        m.C.set_values(C_init_dic)                            
        m.E_ind.set_values(E_ind_init_dic)                    
        m.E_tot.set_values(E_tot_init_dic)                          
        m.M_at.set_values(M_at_init_dic)        
        m.M_up.set_values(M_up_init_dic)                                       
        m.M_lo.set_values(M_lo_init_dic)                                       
        m.T_at.set_values(T_at_init_dic)                                       
        m.T_lo.set_values(T_lo_init_dic)                                       
        m.F.set_values(F_init_dic)                                             
        for pl in m.mC:
            for tt in m.t:
                m.mu[pl, tt].fix(pe.value(m.mu[pl, tt]))
                m.S[pl, tt].fix(pe.value(m.S[pl, tt]))
        coal = {countries[k]: i[k] for k in range(len(i))}
        stop_rule = False
        res_track = []
        count_iter = 0
        # This tolerance relates to the results of the control variable m.mu between
        # successive best response iterations. The number should always be greater (less precision ) 
        # than the provided tol (precision of the optimization algorithm) otherwise 
        # convergence is likely not to be achieved for rounding errors. The difference should
        # be of at least 2 orders of magnitude: e.g. 1e-5 and 1e-7
        tol_2 = 0.00001
        # Change the following number to increase\reduce the max number of iterations to
        # find convergence among successive best response solutions. If the limit is
        # reached, the given solution is likely not to be optimal. This number is necessary to 
        # avoid infinite loops. 
        iter_lim = 25           
        while stop_rule == False and count_iter <= iter_lim:
            res_track.append([])    
            for player in m.mC:
                for tt in m.t:
                    m.mu[player, tt].fixed = False
                    m.S[player, tt].fixed = False
                if coal[player] == 1:
                    obj_expr = sum(m.U[k, j]*coal[k] for k in m.mC for j in m.t)
                else:
                    obj_expr = sum(m.U[player, j] for k in m.mC for j in m.t)
                m.obj = pe.Objective(expr = obj_expr, sense=pe.maximize) 
                opt.solve(m)
                del m.obj 
                res_track[count_iter].append(pe.value(m.mu[player, :]))
                for tt in m.t:
                    m.mu[player, tt].fix(pe.value(m.mu[player, tt]))
                    m.S[player, tt].fix(pe.value(m.S[player, tt]))
            if count_iter != 0:
                # Check on convergence is performed only on m.mu, not on m.S for 
                # shortening computation time
                stop_rule = all(res_track[-1][i][j] >= res_track[-2][i][j] - tol_2 and \
                                res_track[-1][i][j] <= res_track[-2][i][j] + tol_2     \
                                for i in range(len(mC)) for j in range(T-1))            
            count_iter += 1
        coa_n = [countries[j] for j in range(len(countries)) if i[j] == 1]
        coa_res.append([model_res_to_dict(m), coa_n, count_iter])
    
    if coa_c != 'all':
        res_coa = []
        for i in coa_res:
            res_coa.append(output_format(countries, i[0], t, T))
        coa_TU = [[sum(coa_v[i].loc['U',:]) for i in countries] for coa_v in res_coa]
        cc_TU = [coa_TU[i][i] for i in range(len(coa_TU))]
        
        exl_file = load_workbook(results_path+'coop.xlsx')
        sheet = exl_file['global']
        sheet['k12'] = 'Internally Stable'
        sheet['k14'] = 'Externallly Stable'
        sheet['k16'] = 'Fully Stable'
        sheet['k18'] = 'PIS (Potential Internally Stable)'
        
        if all(coop_TU[i] >= cc_TU[i] for i in range(len(coop_TU))):
            sheet['m12'] = "True"
            sheet['m16'] = "True"
        else:
            sheet['m12'] = "False"
            sheet['m16'] = "False"
        sheet['M14'] = 'Not Applicable'
            
        if sum(coop_TU) >= sum(cc_TU):
            sheet['m18'] = "True"
        else:
            sheet['m18'] = "False"
        exl_file.save(filename = results_path+'coop.xlsx')

# Non cooperative result (no coalition being formed)

if nc_c or coa_c == 'all':
    # Resetting starting values (required only if cooperative solution has been computed)
    if coop_c == True:
        m.U.set_values(U_init_dic)                                       
        m.I.set_values(I_init_dic)
        m.S.set_values(S_init_dic)                            
        m.Q.set_values(Q_init_dic)                           
        m.Y.set_values(Y_init_dic)                                       
        m.AB.set_values(AB_init_dic)            
        m.D.set_values(D_init_dic)                            
        m.C.set_values(C_init_dic)                            
        m.E_ind.set_values(E_ind_init_dic)                    
        m.E_tot.set_values(E_tot_init_dic)                          
        m.M_at.set_values(M_at_init_dic)        
        m.M_up.set_values(M_up_init_dic)                                       
        m.M_lo.set_values(M_lo_init_dic)                                       
        m.T_at.set_values(T_at_init_dic)                                       
        m.T_lo.set_values(T_lo_init_dic)                                       
        m.F.set_values(F_init_dic)   
    for pl in m.mC:
        for tt in m.t:
            m.mu[pl, tt].fix(pe.value(m.mu[pl, tt]))
            m.S[pl, tt].fix(pe.value(m.S[pl, tt]))
    stop_rule = False
    res_track = []
    count_iter = 0
    # This tolerance relates to the results of the control variable m.mu between
    # successive best response iterations. The number should always be greater (less precision) 
    # than the provided tol (precision of the optimization algorithm) otherwise 
    # convergence is likely not to be achieved for rounding errors. The difference should
    # be of at least 2 orders of magnitude: e.g. 1e-5 and 1e-7
    tol_2 = 0.00001
    # Change the following number to increase\reduce the max number of iterations to
    # find convergence among successive best response solutions. If the limit is
    # reached, the given solution is likely not to be optimal. This number is necessary to 
    # avoid infinite loops. 
    iter_lim = 25
    while stop_rule == False and count_iter <= iter_lim:
        res_track.append([])    
        for player in m.mC:
            for tt in m.t:
                m.mu[player, tt].fixed = False
                m.S[player, tt].fixed = False
            obj_expr = sum(m.U[player, j] for j in m.t)
            m.obj = pe.Objective(expr = obj_expr, sense=pe.maximize) 
            opt.solve(m)
            del m.obj
            res_track[count_iter].append(pe.value(m.mu[player, :]))
            for tt in m.t:
                m.mu[player, tt].fix(pe.value(m.mu[player, tt]))
                m.S[player, tt].fix(pe.value(m.S[player, tt]))
        if count_iter != 0:
            # Check on convergence is performed only on m.mu, not on m.I for 
            # shortening computation time
            stop_rule = all(res_track[-1][i][j] >= res_track[-2][i][j] - tol_2 and \
                            res_track[-1][i][j] <= res_track[-2][i][j] + tol_2     \
                            for i in range(len(mC)) for j in range(T-1))            
        count_iter += 1
    
    nc_dict = model_res_to_dict(m)
    
    res_nc = output_format(countries, nc_dict, t, T)
    
    results_to_excel(res_nc, countries, results_path, 'non_coop.xlsx')
    book = load_workbook(results_path + 'non_coop.xlsx')
    sheet = book['global']
    sheet['b10'] = 'Number of iterations'
    sheet['b11'] = count_iter
    book.save(filename = results_path + 'non_coop.xlsx')

# All intermediate coalitions (from 2 player coalitions to N-1)

if coa_c != 'none':
    if coa_c == 'all':
        coa = coa_f(len(countries))[len(countries):-1]
    else:
        coa_m = [coa_to_analyse(coa_c)]
        int_coa = coa_int(coa_m[0])
        ext_coa = coa_ext(coa_m[0])
        coa = int_coa.copy()
        coa.append(coa_m[0])
        for i in ext_coa:
            coa.append(i)
        main_coa_id = len(int_coa) 
        members_id = [i for i in range(len(coa_m)) if coa_m[i] == 1]
        non_members_id = [i for i in range(len(coa_m)) if coa_m[i] == 0]
        
    coa_res = []     
                        
    for i in coa:
        m.U.set_values(U_init_dic)                                       
        m.I.set_values(I_init_dic)
        m.S.set_values(S_init_dic)                            
        m.Q.set_values(Q_init_dic)                           
        m.Y.set_values(Y_init_dic)                                       
        m.AB.set_values(AB_init_dic)            
        m.D.set_values(D_init_dic)                            
        m.C.set_values(C_init_dic)                            
        m.E_ind.set_values(E_ind_init_dic)                    
        m.E_tot.set_values(E_tot_init_dic)                          
        m.M_at.set_values(M_at_init_dic)        
        m.M_up.set_values(M_up_init_dic)                                       
        m.M_lo.set_values(M_lo_init_dic)                                       
        m.T_at.set_values(T_at_init_dic)                                       
        m.T_lo.set_values(T_lo_init_dic)                                       
        m.F.set_values(F_init_dic)                                             
        for pl in m.mC:
            for tt in m.t:
                m.mu[pl, tt].fix(pe.value(m.mu[pl, tt]))
                m.S[pl, tt].fix(pe.value(m.S[pl, tt]))
        coac = {countries[k]: i[k] for k in range(len(i))}
        stop_rule = False
        res_track = []
        count_iter = 0
        # This tolerance relates to the results of the control variable m.mu between
        # successive best response iterations. The number should always be greater (less precision ) 
        # than the provided tol (precision of the optimization algorithm) otherwise 
        # convergence is likely not to be achieved for rounding errors. The difference should
        # be of at least 2 orders of magnitude: e.g. 1e-5 and 1e-7
        tol_2 = 0.00001
        # Change the following number to increase\reduce the max number of iterations to
        # find convergence among successive best response solutions. If the limit is
        # reached, the given solution is likely not to be optimal. This number is necessary to 
        # avoid infinite loops. 
        iter_lim = 25           
        while stop_rule == False and count_iter <= iter_lim:
            res_track.append([])    
            for player in m.mC:
                for tt in m.t:
                    m.mu[player, tt].fixed = False
                    m.S[player, tt].fixed = False
                if coac[player] == 1:
                    obj_expr = sum(m.U[k, j]*coac[k] for k in m.mC for j in m.t)
                else:
                    obj_expr = sum(m.U[player, j] for k in m.mC for j in m.t)
                m.obj = pe.Objective(expr = obj_expr, sense=pe.maximize) 
                opt.solve(m)
                del m.obj 
                res_track[count_iter].append(pe.value(m.mu[player, :]))
                for tt in m.t:
                    m.mu[player, tt].fix(pe.value(m.mu[player, tt]))
                    m.S[player, tt].fix(pe.value(m.S[player, tt]))
            if count_iter != 0:
                # Check on convergence is performed only on m.mu, not on m.S for 
                # shortening computation time
                stop_rule = all(res_track[-1][i][j] >= res_track[-2][i][j] - tol_2 and \
                                res_track[-1][i][j] <= res_track[-2][i][j] + tol_2     \
                                for i in range(len(mC)) for j in range(T-1))            
            count_iter += 1
        coa_n = [countries[j] for j in range(len(countries)) if i[j] == 1]
        coa_res.append([model_res_to_dict(m), coa_n, count_iter])
    
    # Saving results of coalitions in excell files (one for each coalition)

    if coa_c == 'all':    
        ccc = 1
        res_coa = []
        for i in coa_res:
            res_coa.append(output_format(countries, i[0], t, T))
            file_name = 'coa'+str(ccc)+'.xlsx'
            results_to_excel(res_coa[-1], countries, results_path, file_name)
            book = load_workbook(results_path + 'coa' + str(ccc) + '.xlsx')
            book.create_sheet(title='Members')
            new_sheet = book['Members']
            new_sheet['b2'] = 'Coalition Members'
            new_sheet['b3'] = str(i[1])
            new_sheet['b5'] = 'Number of iterations'
            new_sheet['b6'] = str(i[2])
            book.save(filename = results_path + 'coa' + str(ccc) + '.xlsx')
            ccc += 1
  
        pay = []
        for i in range(1,int(2**N)-(N+1)):       
            xls = pd.ExcelFile(results_path+'coa'+str(i)+'.xlsx')
            pay.append([])
            U_c = []
            for i in countries:
                U_c.append(pd.read_excel(xls, i))
                U_c[-1].set_index('Unnamed: 0', inplace=True)
                pay[-1].append(float(U_c[-1].loc['U'].sum()))
        
        xls = pd.ExcelFile(results_path+'non_coop'+'.xlsx')
        U_c = []
        p_nc = []
        for i in countries:
            U_c.append(pd.read_excel(xls, i))
            U_c[-1].set_index('Unnamed: 0', inplace=True)
            p_nc.append(float(U_c[-1].loc['U'].sum()))
        for i in range(N):
            pay.insert(0, p_nc)
        
        xls = pd.ExcelFile(results_path+'non_coop'+'.xlsx')
        U_c = []
        p_coop = []
        for i in countries:
            U_c.append(pd.read_excel(xls, i))
            U_c[-1].set_index('Unnamed: 0', inplace=True)
            p_coop.append(float(U_c[-1].loc['U'].sum()))
        pay.append(p_coop)
        
            
        sps = [[1 if i == j else 0 for i in range(N)] for j in range(N)]
        gr_coa = [1 for i in range(N)]
        
        coa_s = sps
        for i in coa:
            coa_s.append(i)
        coa_s.append(gr_coa)
        
        int_diff = c_f_dif(coa_s, pay)
        
        int_c = int_st(int_diff, coa_s)
        
        ext_diff = dif_ext(coa_s, pay)
        
        ext_c = ext_st(ext_diff, coa_s)
        
        full_st = stab_c(coa_s, int_c[0], ext_c[0])
        
        full_st = [i for i in full_st if i != 0]
        
        #Write results on excel files
        
        # Cooperative case
        exl_file = load_workbook(results_path+'coop.xlsx')
        sheet = exl_file['global']
        sheet['k12'] = 'Internally Stable'
        sheet['k14'] = 'Externallly Stable'
        sheet['k16'] = 'Fully Stable'
        sheet['k18'] = 'PIS (Potential Internally Stable)'
        
        if gr_coa in int_c[0]:
            sheet['m12'] = "True"
        else:
            sheet['m12'] = "False"
        sheet['M14'] = 'Not Applicable'
        
        if gr_coa in full_st:
            sheet['m16'] = "True"
        else:
            sheet['m16'] = "False"
        
        if gr_coa in int_c[1]:
            sheet['m18'] = "True"
        else:
            sheet['m18'] = "False"
        exl_file.save(filename = results_path+'coop.xlsx')
        
        # Other coalitions        
        for i in range(len(coa_s[N:-1])):
            exl_file = load_workbook(results_path+'coa'+str(i+1)+'.xlsx')
            sheet = exl_file['global']
            sheet['k12'] = 'Internally Stable'
            sheet['k14'] = 'Externallly Stable'
            sheet['k16'] = 'Fully Stable'
            sheet['k18'] = 'PIS (Potential Internally Stable)'
            
            if coa_s[i] in int_c[0]:
                sheet['m12'] = "True"
            else:
                sheet['m12'] = "False"
        
            if coa_s[i] in ext_c[0]:
                sheet['m14'] = "True"
            else:
                sheet['m14'] = "False"
            
            if coa_s[i] in full_st:
                sheet['m16'] = "True"
            else:
                sheet['m16'] = "False"
            
            if coa_s[i] in int_c[1] or coa_s[i] in int_c[0]:
                sheet['m18'] = "True"
            else:
                sheet['m18'] = "False"
                
            exl_file.save(filename = results_path+'coa'+str(i+1)+'.xlsx')
    
    else:
        res_coa = []
        for i in coa_res:
            res_coa.append(output_format(countries, i[0], t, T))
        
        file_name = coa_c+'.xlsx'
        results_to_excel(res_coa[-1], countries, results_path, file_name)
        book = load_workbook(results_path + coa_c + '.xlsx')
        book.create_sheet(title='Members')
        new_sheet = book['Members']
        new_sheet['b2'] = 'Coalition Members'
        new_sheet['b3'] = str(i[1])
        new_sheet['b5'] = 'Number of iterations'
        new_sheet['b6'] = str(i[2])
        book.save(filename = results_path + coa_c + '.xlsx')   
        
        ac_TU = [[sum(coa_v[i].loc['U',:]) for i in countries] for coa_v in res_coa]
        int_c_TU = [ac_TU[i][members_id[i]] for i in range(len(members_id))]
        m_c_TU = [ac_TU[main_coa_id][i] for i in members_id]
        m_nc_TU = [ac_TU[main_coa_id][i] for i in non_members_id]
        ext_c_TU = [ac_TU[i][non_members_id[i]] for i in range(len(non_members_id))]
                   
        exl_file = load_workbook(results_path + coa_c + '.xlsx')
        sheet = exl_file['global']
        sheet['k12'] = 'Internally Stable'
        sheet['k14'] = 'Externallly Stable'
        sheet['k16'] = 'Fully Stable'
        sheet['k18'] = 'PIS (Potential Internally Stable)'
        full_1 = False
        full_2 = False
        
        if all(m_c_TU[i] >= int_c_TU[i] for i in range(len(m_c_TU))):
            sheet['m12'] = "True"
            full_1 = True
        else:
            sheet['m12'] = "False"

        if all(ext_c_TU[i] >= m_nc_TU[i] for i in range(len(m_nc_TU))):
            sheet['m14'] = "True"
            full_2 = True
        else:
            sheet['m14'] = "False"
            
        if full_1 and full_2:
            sheet['m16'] = "True"
        else:
            sheet['m16'] = "False"            
            
        if sum(m_c_TU) >= sum(int_c_TU):
            sheet['m18'] = "True"
        else:
            sheet['m18'] = "False"
        exl_file.save(filename = results_path + coa_c + '.xlsx')        
    
    
    
    
    
    
    
       
    
